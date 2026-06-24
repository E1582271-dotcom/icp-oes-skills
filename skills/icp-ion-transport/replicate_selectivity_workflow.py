# -*- coding: utf-8 -*-
"""
replicate_selectivity_workflow  (reconstructed 2026-06-15)
==========================================================
Rebuilds the 8-sheet Results.xlsx from a ``1_Original Data`` sheet using the
replicate-first methodology, and provides the selectivity-figure helper.

Reconstructed faithfully from the existing Results.xlsx outputs (verified to 8
decimals against Current Density / New membrane results).  Key methodology:

  * Sampling correction is ADDITIVE / cumulative (NOT the geometric formula that
    was mistakenly written in the old sheet notes):
        C_corr(t_i) = C_meas(t_i) + (Vs/V) * sum_{j<i} C_meas(t_j)
    where the sum runs over every *measured* prior time point (so the real
    sampling sequence is preserved even when only a subset of time points is
    exported).
  * Replicate-first: every replicate cell is corrected / fluxed / made selective
    independently; Mean = average over replicates.  Works for any number of
    replicates per group (2 for a single run, 4 for two pooled parallel runs).
  * Selectivity NQ filter: a replicate is "Not Quantified" and dropped from the
    mean when JCa_eff <= 0, JNa_eff <= 0, or |JNa_eff| < 1e-6.
  * Mass balance (CON+DIL) and current efficiency at the final exported time.

Entry points used by the calculate_results.py / plot_figure3.py wrappers:
    rebuild_results_workbook(xlsx_path, default_current_density=None,
                             time_points=None)
    plot_selectivity_figures(xlsx_path, out_dir, mode="current_density")
"""

from __future__ import annotations

import math
import re
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ----------------------------------------------------------------------------
# Physical / experimental parameters
# ----------------------------------------------------------------------------
V_mL = 20.0
V_sample_mL = 0.4          # 2 x 0.2 mL, replenished
A_cm2 = 1.54
M_Ca = 40.08
M_Na = 22.99
DILUTION = 75
F_CONST = 96485.33212      # C/mol

V_L = V_mL / 1000.0
RATIO = V_sample_mL / V_mL  # Vs/V for the additive correction
NQ = "NQ"
EM = "—"              # em dash

PARAMS_TXT = (
    f"Parameters: V = {V_mL} mL | V_sample = {V_sample_mL} mL (2x0.2 mL, replenished) | "
    f"A = {A_cm2} cm² | M(Ca) = {M_Ca} g/mol | M(Na) = {M_Na} g/mol | Dilution = {DILUTION}x"
)

# ----------------------------------------------------------------------------
# Styling helpers
# ----------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
SEC_FILL = PatternFill("solid", fgColor="D6E4F0")
SEC_FONT = Font(name="Arial", bold=True, color="1A3A5C", size=9)
DAT_FONT = Font(name="Arial", size=9)
UNIT_FONT = Font(name="Arial", italic=True, size=8, color="777777")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
CTR_NW = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center")
_THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_GROUP_PALETTE = [
    "C0392B", "27AE60", "8E44AD", "1A6496", "D35400",
    "16A085", "2C3E50", "7F8C8D", "C2185B", "00838F",
]


def _hdr(ws, r, c, text, span=1, fill=None, font=None, wrap=True):
    cell = ws.cell(row=r, column=c, value=text)
    cell.fill = fill or HDR_FILL
    cell.font = font or HDR_FONT
    cell.alignment = CTR if wrap else CTR_NW
    cell.border = BORDER
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    return cell


def _dat(ws, r, c, val, fmt="0.0000"):
    if val is None:
        cell = ws.cell(row=r, column=c, value=EM)
        cell.alignment = CTR_NW
    elif isinstance(val, str):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = CTR_NW
    else:
        cell = ws.cell(row=r, column=c, value=round(float(val), 8))
        cell.number_format = fmt
        cell.alignment = CTR_NW
    cell.font = DAT_FONT
    cell.border = BORDER
    return cell


def _note(ws, r, c, text, span):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = UNIT_FONT
    cell.alignment = LFT
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)


def _no_grid(ws):
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------------------
# Reading + computation
# ----------------------------------------------------------------------------
def _to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", EM, "####", "Uncal", "-", "nan", "NQ"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_current_density(group_name, default):
    """Return current density (mA cm^-2) parsed from a group label."""
    m = re.search(r"(\d+(?:\.\d+)?)\s*mA", str(group_name))
    if m:
        return float(m.group(1))
    return default


def read_original(ws):
    """Read the ``1_Original Data`` sheet into an ordered structure.

    Returns OrderedDict: group -> {"CON": {rep: {time: (ca_ppm, na_ppm)}},
                                    "DIL": {...}}  preserving first-seen order.
    """
    rows = list(ws.iter_rows(values_only=True))
    data = OrderedDict()
    for r in rows[1:]:
        if not r or r[0] in (None, ""):
            continue
        group = str(r[0]).strip()
        side = str(r[1]).strip() if r[1] is not None else ""
        rep = r[2]
        t = r[3]
        ca_ppm = _to_num(r[6])
        na_ppm = _to_num(r[8])
        if side not in ("CON", "DIL") or rep is None or t is None:
            continue
        rep = int(rep)
        t = int(round(float(t)))
        g = data.setdefault(group, {"CON": OrderedDict(), "DIL": OrderedDict()})
        g[side].setdefault(rep, {})[t] = (ca_ppm, na_ppm)
    return data


def _corrected_series(times, conc, ratio=RATIO):
    """Additive cumulative sampling correction over the full measured series.

    ``ratio`` = V_sample/V for this replicate (defaults to the global RATIO).
    A per-replicate ratio lets arms sampled with different V_sample (e.g.
    0.1 vs 0.4 mL) each be corrected with their own value before merging.
    """
    corr = []
    running = 0.0
    for i, c in enumerate(conc):
        if c is None:
            corr.append(None)
            continue
        corr.append(c + ratio * running)
        running += c
    return corr


def _mean(vals):
    xs = [v for v in vals if v is not None]
    return sum(xs) / len(xs) if xs else None


def compute_group(side_data, all_times, out_times, vs_map=None):
    """Compute per-replicate corrected conc, dC, flux, plus means.

    Returns a dict keyed by metric with structure:
      reps -> sorted list of rep ids
      'ca_corr'[rep][t], 'na_corr', 'dca', 'dna' (mmol/L)
      'jca'[rep][t], 'jna'  (mmol/cm2/min)
      '*_mean'[t] for ca_corr/na_corr/dca/dna/jca/jna

    ``vs_map`` optionally maps rep id -> V_sample (mL) so replicates sampled
    with different volumes are each corrected with their own ratio.  Reps not
    present fall back to the global V_sample_mL.
    """
    reps = sorted(side_data.keys())
    res = {"reps": reps}
    for key in ("ca_corr", "na_corr", "dca", "dna", "jca", "jna"):
        res[key] = {rep: {} for rep in reps}

    for rep in reps:
        series = side_data[rep]
        ca = [(series.get(t, (None, None))[0]) for t in all_times]
        na = [(series.get(t, (None, None))[1]) for t in all_times]
        ca_mmol = [None if c is None else c / M_Ca for c in ca]
        na_mmol = [None if c is None else c / M_Na for c in na]
        ratio = (vs_map.get(rep, V_sample_mL) if vs_map else V_sample_mL) / V_mL
        ca_corr = _corrected_series(all_times, ca_mmol, ratio)
        na_corr = _corrected_series(all_times, na_mmol, ratio)
        c0_ca = ca_corr[0] if ca_corr and ca_corr[0] is not None else 0.0
        c0_na = na_corr[0] if na_corr and na_corr[0] is not None else 0.0
        for i, t in enumerate(all_times):
            cac, nac = ca_corr[i], na_corr[i]
            res["ca_corr"][rep][t] = cac
            res["na_corr"][rep][t] = nac
            dca = None if cac is None else cac - c0_ca
            dna = None if nac is None else nac - c0_na
            res["dca"][rep][t] = dca
            res["dna"][rep][t] = dna
            if t == 0:
                res["jca"][rep][t] = None
                res["jna"][rep][t] = None
            else:
                res["jca"][rep][t] = None if dca is None else dca * V_L / A_cm2 / t
                res["jna"][rep][t] = None if dna is None else dna * V_L / A_cm2 / t

    for key in ("ca_corr", "na_corr", "dca", "dna", "jca", "jna"):
        res[key + "_mean"] = {t: _mean([res[key][rep].get(t) for rep in reps]) for t in out_times}
    return res


def _selectivity_rows(reps, jca, jna, out_times, removal=False):
    """Per-rep selectivity with NQ filter. removal=True -> DIL side (use -J)."""
    out = {}
    for t in out_times:
        if t == 0:
            out[t] = {"jca_mean": None, "jna_mean": None, "per_rep": {r: EM for r in reps},
                      "mean": None, "valid": None, "excluded": EM}
            continue
        jca_eff_mean = _mean([(-jca[r].get(t) if (removal and jca[r].get(t) is not None) else jca[r].get(t))
                              for r in reps])
        jna_eff_mean = _mean([(-jna[r].get(t) if (removal and jna[r].get(t) is not None) else jna[r].get(t))
                              for r in reps])
        per_rep = {}
        sels = []
        excluded = []
        for r in reps:
            jc = jca[r].get(t)
            jn = jna[r].get(t)
            if jc is None or jn is None:
                per_rep[r] = NQ
                excluded.append(r)
                continue
            jc_eff = -jc if removal else jc
            jn_eff = -jn if removal else jn
            if jc_eff <= 0 or jn_eff <= 0 or abs(jn_eff) < 1e-6:
                per_rep[r] = NQ
                excluded.append(r)
                continue
            s = jc_eff / jn_eff
            per_rep[r] = s
            sels.append(s)
        out[t] = {
            "jca_mean": jca_eff_mean,
            "jna_mean": jna_eff_mean,
            "per_rep": per_rep,
            "mean": (sum(sels) / len(sels)) if sels else None,
            "valid": len(sels) if sels else 0,
            "excluded": ", ".join(str(r) for r in excluded) if excluded else EM,
        }
    return out


# ----------------------------------------------------------------------------
# Workbook writer
# ----------------------------------------------------------------------------
def _conc_sheet(wb, title, groups, comp, side, out_times):
    ws = wb.create_sheet(title)
    _no_grid(ws)
    names = list(groups)
    # 4 metrics each (nrep+1) cols, + gap
    def width(n):
        nrep = len(comp[n][side]["reps"]) or 1
        return (nrep + 1) * 4
    starts = {}
    c = 2
    for n in names:
        starts[n] = c
        c += width(n) + 1
    total = c - 2

    label = "Concentrate Cell (CON side)" if side == "CON" else "Dilute Cell (DI side)"
    _hdr(ws, 1, 1, f"Concentration in {label} [mmol/L, sampling-corrected]", span=total + 1,
         fill=SEC_FILL, font=SEC_FONT)
    _note(ws, 2, 1, PARAMS_TXT, total + 1)
    _note(ws, 3, 1, "Replicate-first: each replicate corrected independently (additive cumulative "
                    "sampling correction); Mean = average of replicates and is the plotted point.", total + 1)

    _hdr(ws, 4, 1, "")
    for gi, n in enumerate(names):
        gc = PatternFill("solid", fgColor=_GROUP_PALETTE[gi % len(_GROUP_PALETTE)])
        _hdr(ws, 4, starts[n], n, span=width(n), fill=gc)

    _hdr(ws, 5, 1, "Time (min)")
    for n in names:
        nrep = len(comp[n][side]["reps"]) or 1
        reps = comp[n][side]["reps"]
        cs = starts[n]
        for mi, mlabel in enumerate(("Ca2+ corr", "Na+ corr", "dCa2+", "dNa+")):
            base = cs + mi * (nrep + 1)
            for ri, rep in enumerate(reps):
                _hdr(ws, 5, base + ri, f"{mlabel}\nRep-{rep}")
            _hdr(ws, 5, base + nrep, f"{mlabel}\nMean")
    ws.row_dimensions[5].height = 30

    for i, t in enumerate(out_times):
        r = 6 + i
        _dat(ws, r, 1, t, fmt="0")
        for n in names:
            d = comp[n][side]
            reps = d["reps"]
            nrep = len(reps) or 1
            cs = starts[n]
            metrics = [("ca_corr", "ca_corr_mean"), ("na_corr", "na_corr_mean"),
                       ("dca", "dca_mean"), ("dna", "dna_mean")]
            for mi, (mk, mm) in enumerate(metrics):
                base = cs + mi * (nrep + 1)
                for ri, rep in enumerate(reps):
                    _dat(ws, r, base + ri, d[mk][rep].get(t))
                _dat(ws, r, base + nrep, d[mm].get(t))
    return ws


def _flux_sheet(wb, title, groups, comp, side, out_times):
    ws = wb.create_sheet(title)
    _no_grid(ws)
    names = list(groups)

    def width(n):
        nrep = len(comp[n][side]["reps"]) or 1
        return (nrep + 1) * 2
    starts = {}
    c = 2
    for n in names:
        starts[n] = c
        c += width(n) + 1
    total = c - 2

    label = "CON side" if side == "CON" else "DI side"
    _hdr(ws, 1, 1, f"Ion Flux ({label}) [mmol/cm²/min]", span=total + 1, fill=SEC_FILL, font=SEC_FONT)
    _note(ws, 2, 1, PARAMS_TXT, total + 1)
    _note(ws, 3, 1, "Replicate-first: J = dC_corr * V / A / t per replicate; Mean = average of replicates.",
          total + 1)

    _hdr(ws, 4, 1, "")
    for gi, n in enumerate(names):
        gc = PatternFill("solid", fgColor=_GROUP_PALETTE[gi % len(_GROUP_PALETTE)])
        _hdr(ws, 4, starts[n], n, span=width(n), fill=gc)

    _hdr(ws, 5, 1, "Time (min)")
    for n in names:
        reps = comp[n][side]["reps"]
        nrep = len(reps) or 1
        cs = starts[n]
        for mi, mlabel in enumerate(("J Ca2+", "J Na+")):
            base = cs + mi * (nrep + 1)
            for ri, rep in enumerate(reps):
                _hdr(ws, 5, base + ri, f"{mlabel} Rep-{rep}\n(mmol/cm²/min)")
            _hdr(ws, 5, base + nrep, f"{mlabel} Mean\n(mmol/cm²/min)")
    ws.row_dimensions[5].height = 30

    for i, t in enumerate(out_times):
        r = 6 + i
        _dat(ws, r, 1, t, fmt="0")
        for n in names:
            d = comp[n][side]
            reps = d["reps"]
            nrep = len(reps) or 1
            cs = starts[n]
            for mi, (mk, mm) in enumerate([("jca", "jca_mean"), ("jna", "jna_mean")]):
                base = cs + mi * (nrep + 1)
                for ri, rep in enumerate(reps):
                    _dat(ws, r, base + ri, d[mk][rep].get(t), fmt="0.000000")
                _dat(ws, r, base + nrep, d[mm].get(t), fmt="0.000000")
    return ws


def _selectivity_sheet(wb, groups, comp, out_times):
    ws = wb.create_sheet("6_Selectivity")
    _no_grid(ws)
    names = list(groups)

    def width(n):
        nrep = len(comp[n]["CON"]["reps"]) or 1
        return 2 + nrep + 3   # JCa, JNa, sel per rep, Mean, Valid, Excluded
    starts = {}
    c = 2
    for n in names:
        starts[n] = c
        c += width(n) + 1
    total = c - 2

    _hdr(ws, 1, 1, "Permselectivity S(Ca2+/Na+) (CON-based and DIL-based, replicate-first)",
         span=total + 1, fill=SEC_FILL, font=SEC_FONT)
    _note(ws, 2, 1, PARAMS_TXT, total + 1)
    _note(ws, 3, 1, "NQ = Not quantified and excluded from Mean. A replicate is marked NQ when "
                    "JCa_eff <= 0, JNa_eff <= 0, or |JNa_eff| < 1e-6. Valid reps used for mean = number "
                    "of replicate selectivity values included in Mean.", total + 1)

    def section(top, side, removal, title, flux_label):
        _hdr(ws, top, 1, title, span=total + 1, fill=SEC_FILL, font=SEC_FONT)
        _hdr(ws, top + 1, 1, "")
        for gi, n in enumerate(names):
            gc = PatternFill("solid", fgColor=_GROUP_PALETTE[gi % len(_GROUP_PALETTE)])
            _hdr(ws, top + 1, starts[n], n, span=width(n), fill=gc)
        _hdr(ws, top + 2, 1, "Time (min)")
        for n in names:
            reps = comp[n][side]["reps"]
            cs = starts[n]
            _hdr(ws, top + 2, cs, f"J Ca2+ {flux_label}\n(mmol/cm²/min)")
            _hdr(ws, top + 2, cs + 1, f"J Na+ {flux_label}\n(mmol/cm²/min)")
            for ri, rep in enumerate(reps):
                _hdr(ws, top + 2, cs + 2 + ri, f"Selectivity from\nRep-{rep}")
            nrep = len(reps) or 1
            _hdr(ws, top + 2, cs + 2 + nrep, "Mean selectivity\n(valid reps only)")
            _hdr(ws, top + 2, cs + 3 + nrep, "Valid reps\nused for mean")
            _hdr(ws, top + 2, cs + 4 + nrep, "Excluded reps\n(NQ)")
        ws.row_dimensions[top + 2].height = 30
        for i, t in enumerate(out_times):
            r = top + 3 + i
            _dat(ws, r, 1, t, fmt="0")
            for n in names:
                reps = comp[n][side]["reps"]
                nrep = len(reps) or 1
                cs = starts[n]
                selrows = comp[n][side + "_sel"]
                row = selrows[t]
                _dat(ws, r, cs, row["jca_mean"], fmt="0.000000")
                _dat(ws, r, cs + 1, row["jna_mean"], fmt="0.000000")
                for ri, rep in enumerate(reps):
                    pr = row["per_rep"].get(rep)
                    _dat(ws, r, cs + 2 + ri, pr, fmt="0.0000")
                _dat(ws, r, cs + 2 + nrep, row["mean"], fmt="0.0000")
                _dat(ws, r, cs + 3 + nrep, (row["valid"] if row["valid"] is not None else EM), fmt="0")
                _dat(ws, r, cs + 4 + nrep, row["excluded"])
        return top + 3 + len(out_times)

    after = section(5, "CON", False, "CON-based selectivity", "CON")
    section(after + 2, "DIL", True, "DIL-based selectivity", "removal")
    return ws


def _group_final_time(comp_group, out_times):
    """Largest exported non-zero time where this group has CON & DIL conc means."""
    con, dil = comp_group["CON"], comp_group["DIL"]
    cands = [t for t in out_times if t != 0
             and con["ca_corr_mean"].get(t) is not None
             and dil["ca_corr_mean"].get(t) is not None]
    if cands:
        return max(cands)
    cands = [t for t in out_times if t != 0 and con["ca_corr_mean"].get(t) is not None]
    return max(cands) if cands else max(t for t in out_times if t != 0)


def _mass_balance_sheet(wb, groups, comp, out_times):
    ws = wb.create_sheet("7_Mass Balance")
    _no_grid(ws)
    _hdr(ws, 1, 1, "Final-point Mass Balance (CON + DIL)", span=12, fill=SEC_FILL, font=SEC_FONT)
    _note(ws, 2, 1, PARAMS_TXT, 12)
    _note(ws, 3, 1, "Uses final corrected concentration means. Selectivity NQ filtering is not applied "
                    "to mass balance.", 12)
    headers = ["Group", "Final time (min)", "Ca total at 0 min (mmol)", "Ca total at final min (mmol)",
               "Ca recovery (%)", "Ca error (mmol)", "Na total at 0 min (mmol)",
               "Na total at final min (mmol)", "Na recovery (%)", "Na error (mmol)",
               "Overall molar cation recovery (%)", "Overall charge recovery (%)"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 4, c, h)
    for gi, n in enumerate(groups):
        r = 5 + gi
        con, dil = comp[n]["CON"], comp[n]["DIL"]
        final_t = _group_final_time(comp[n], out_times)

        def total(metric_mean, t):
            cc = con[metric_mean].get(t)
            dd = dil[metric_mean].get(t)
            if cc is None or dd is None:
                return None
            return (cc + dd) * V_L

        ca0 = total("ca_corr_mean", 0)
        caf = total("ca_corr_mean", final_t)
        na0 = total("na_corr_mean", 0)
        naf = total("na_corr_mean", final_t)
        ca_rec = (caf / ca0 * 100) if (ca0 not in (None, 0) and caf is not None) else None
        na_rec = (naf / na0 * 100) if (na0 not in (None, 0) and naf is not None) else None
        ca_err = (caf - ca0) if (ca0 is not None and caf is not None) else None
        na_err = (naf - na0) if (na0 is not None and naf is not None) else None
        tot0 = (ca0 or 0) + (na0 or 0)
        totf = (caf or 0) + (naf or 0)
        molar_rec = (totf / tot0 * 100) if tot0 else None
        chg0 = 2 * (ca0 or 0) + (na0 or 0)
        chgf = 2 * (caf or 0) + (naf or 0)
        charge_rec = (chgf / chg0 * 100) if chg0 else None
        vals = [n, final_t, ca0, caf, ca_rec, ca_err, na0, naf, na_rec, na_err, molar_rec, charge_rec]
        for c, v in enumerate(vals, 1):
            if c == 1:
                _dat(ws, r, c, v)
            elif c == 2:
                _dat(ws, r, c, v, fmt="0")
            else:
                _dat(ws, r, c, v, fmt="0.000")
    return ws


def _current_efficiency_sheet(wb, groups, comp, out_times, default_cd):
    ws = wb.create_sheet("8_Current Efficiency")
    _no_grid(ws)
    _hdr(ws, 1, 1, "Formula: CON CE_Ca(%) = 2*F*(dCa_CON*V/1000)/(I*t)*100; CON CE_Na(%) = "
                   "F*(dNa_CON*V/1000)/(I*t)*100; DIL CE uses removed amount = -dC_DIL*V/1000. "
                   "Selectivity NQ filtering is not applied.", span=18, fill=SEC_FILL, font=SEC_FONT)
    _note(ws, 2, 1, PARAMS_TXT + f" | F = {F_CONST} C/mol", 18)
    _note(ws, 3, 1, "Final-point current efficiency uses corrected dC means at the final exported time.", 18)
    headers = ["Group", "Final time (min)", "Current density (mA cm^-2)", "Current (mA)",
               "CON dCa2+ (mmol/L)", "CON dNa+ (mmol/L)", "CON transferred Ca2+ (mol)",
               "CON transferred Na+ (mol)", "CON CE Ca2+ (%)", "CON CE Na+ (%)", "CON CE total cation (%)",
               "DIL dCa2+ (mmol/L)", "DIL dNa+ (mmol/L)", "DIL removed Ca2+ (mol)", "DIL removed Na+ (mol)",
               "DIL CE Ca2+ (%)", "DIL CE Na+ (%)", "DIL CE total cation (%)"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 4, c, h)
    for gi, n in enumerate(groups):
        r = 5 + gi
        final_t = _group_final_time(comp[n], out_times)
        cd = parse_current_density(n, default_cd)
        I_mA = cd * A_cm2
        I_A = I_mA / 1000.0
        t_s = final_t * 60.0
        con, dil = comp[n]["CON"], comp[n]["DIL"]
        dca_con = con["dca_mean"].get(final_t)
        dna_con = con["dna_mean"].get(final_t)
        dca_dil = dil["dca_mean"].get(final_t)
        dna_dil = dil["dna_mean"].get(final_t)

        def ce(dc, z, removed=False):
            if dc is None or I_A == 0 or t_s == 0:
                return None, None
            amt = ((-dc if removed else dc) * V_mL / 1000.0) / 1000.0  # mmol/L*mL -> mol
            return amt, z * F_CONST * amt / (I_A * t_s) * 100

        ca_amt_con, ce_ca_con = ce(dca_con, 2)
        na_amt_con, ce_na_con = ce(dna_con, 1)
        ca_amt_dil, ce_ca_dil = ce(dca_dil, 2, removed=True)
        na_amt_dil, ce_na_dil = ce(dna_dil, 1, removed=True)
        ce_tot_con = (ce_ca_con or 0) + (ce_na_con or 0) if ce_ca_con is not None else None
        ce_tot_dil = (ce_ca_dil or 0) + (ce_na_dil or 0) if ce_ca_dil is not None else None
        vals = [n, final_t, cd, I_mA, dca_con, dna_con, ca_amt_con, na_amt_con, ce_ca_con, ce_na_con,
                ce_tot_con, dca_dil, dna_dil, ca_amt_dil, na_amt_dil, ce_ca_dil, ce_na_dil, ce_tot_dil]
        for c, v in enumerate(vals, 1):
            if c == 1:
                _dat(ws, r, c, v)
            elif c in (2, 3):
                _dat(ws, r, c, v, fmt="0.###")
            elif c in (7, 8, 14, 15):
                _dat(ws, r, c, v, fmt="0.000000")
            else:
                _dat(ws, r, c, v, fmt="0.000")
    return ws


def rebuild_results_workbook(xlsx_path, default_current_density=None, time_points=None,
                             exclude=None, vs_overrides=None):
    """Rebuild sheets 2-8 of an existing Results.xlsx in place.

    The workbook must already contain ``1_Original Data`` (and optionally
    ``Calibration``); both are preserved.  ``time_points`` limits which time
    rows are *exported* (the additive correction still uses every measured
    time).  Default: every time present in the data.

    ``exclude`` is an optional set of ``(group, time)`` tuples.  Those measured
    concentrations are treated as MISSING for every side/replicate: they are
    dropped from the additive cumulative correction and never exported.  The raw
    values stay in the ``1_Original Data`` sheet for traceability.  Used to
    discard a confirmed bad measurement (e.g. a dilution-error time point)
    without fabricating a replacement.

    ``vs_overrides`` maps ``(group, rep)`` -> V_sample (mL) for replicates that
    were sampled with a non-default volume; each such replicate's additive
    correction uses its own ratio.  Reps not listed use the global V_sample_mL.
    """
    import os
    xlsx_path = str(xlsx_path)
    exclude = set(exclude or ())
    wb = openpyxl.load_workbook(xlsx_path)
    keep = [s for s in wb.sheetnames if s in ("1_Original Data", "Calibration")]
    data = read_original(wb["1_Original Data"])

    # null out excluded (group, time) points across all sides / replicates
    for group, side_dict in data.items():
        for side in ("CON", "DIL"):
            for rep, tv in side_dict[side].items():
                for t in list(tv):
                    if (group, t) in exclude:
                        tv[t] = (None, None)

    # remove all derived sheets, keep originals
    for s in list(wb.sheetnames):
        if s not in keep:
            del wb[s]

    groups = list(data.keys())
    # determine measured times (union over all reps/sides) and output times
    all_times = sorted({t for g in data.values() for side in ("CON", "DIL")
                        for rep in g[side].values() for t in rep})
    if time_points is None:
        out_times = list(all_times)
    else:
        out_times = [t for t in time_points if t in all_times]

    vs_overrides = dict(vs_overrides or {})
    comp = {}
    for n in groups:
        comp[n] = {}
        for side in ("CON", "DIL"):
            reps_here = sorted(data[n][side].keys())
            vs_map = {rep: vs_overrides.get((n, rep), V_sample_mL) for rep in reps_here}
            comp[n][side] = compute_group(data[n][side], all_times, out_times, vs_map)
        comp[n]["CON_sel"] = _selectivity_rows(
            comp[n]["CON"]["reps"], comp[n]["CON"]["jca"], comp[n]["CON"]["jna"], out_times, removal=False)
        comp[n]["DIL_sel"] = _selectivity_rows(
            comp[n]["DIL"]["reps"], comp[n]["DIL"]["jca"], comp[n]["DIL"]["jna"], out_times, removal=True)
        # attach to side dict for sheet writer convenience
        comp[n]["CON"]["CON_sel"] = comp[n]["CON_sel"]
        comp[n]["DIL"]["DIL_sel"] = comp[n]["DIL_sel"]

    _conc_sheet(wb, "2_CON Concentration", groups, comp, "CON", out_times)
    _conc_sheet(wb, "3_DI Concentration", groups, comp, "DIL", out_times)
    _flux_sheet(wb, "4_CON Ion Flux", groups, comp, "CON", out_times)
    _flux_sheet(wb, "5_DI Ion Flux", groups, comp, "DIL", out_times)
    _selectivity_sheet(wb, groups, comp, out_times)
    _mass_balance_sheet(wb, groups, comp, out_times)
    _current_efficiency_sheet(wb, groups, comp, out_times, default_current_density)

    # order: 1_Original Data, Calibration, then 2..8
    desired = ["1_Original Data", "Calibration", "2_CON Concentration", "3_DI Concentration",
               "4_CON Ion Flux", "5_DI Ion Flux", "6_Selectivity", "7_Mass Balance",
               "8_Current Efficiency"]
    order = [s for s in desired if s in wb.sheetnames]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(xlsx_path)
    return {"workbook": xlsx_path, "groups": groups, "time_points": out_times}


# ----------------------------------------------------------------------------
# Figure 3 — selectivity
# ----------------------------------------------------------------------------
def plot_selectivity_figures(xlsx_path, out_dir, mode="current_density",
                             doping_bases=("CON", "DIL")):
    import os
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams.update({
        "font.family": "sans-serif", "font.sans-serif": ["Arial", "DejaVu Sans"],
        "font.size": 8, "axes.linewidth": 0.8, "pdf.fonttype": 42, "ps.fonttype": 42,
        "savefig.dpi": 600,
    })
    _NPG = ["#E64B35", "#4DBBD5", "#00A087", "#3C5488", "#F39B7F", "#8491B4"]

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["6_Selectivity"]
    rows = list(ws.iter_rows(values_only=True))

    def parse_section(section_label):
        """Parse one selectivity block (CON- or DIL-based) from sheet 6.

        Returns names, mean_cols, rep_col_lists, data_rows, times, final_idx.
        """
        sec_row = next(i for i, r in enumerate(rows) if r and r[0] == section_label)
        band = rows[sec_row + 1]
        header = rows[sec_row + 2]
        names, starts = [], []
        for ci, v in enumerate(band):
            if ci >= 1 and v not in (None, ""):
                names.append(str(v))
                starts.append(ci)
        bounds = starts + [len(band)]
        def mean_col(cs, nxt):
            for ci in range(cs, nxt):
                h = header[ci]
                if h and "Mean selectivity" in str(h):
                    return ci
            return cs + 4
        def rep_cols(cs, nxt):
            return [ci for ci in range(cs, nxt)
                    if header[ci] and "Selectivity from" in str(header[ci])]
        mean_cols = [mean_col(starts[i], bounds[i + 1]) for i in range(len(starts))]
        rep_col_lists = [rep_cols(starts[i], bounds[i + 1]) for i in range(len(starts))]
        data_rows = []
        for r in rows[sec_row + 3:]:
            if r and isinstance(r[0], (int, float)):
                data_rows.append(r)
            elif r and r[0] not in (None, "") and not isinstance(r[0], (int, float)):
                break
        times = [int(r[0]) for r in data_rows]
        return names, mean_cols, rep_col_lists, data_rows, times, len(data_rows) - 1

    def rep_std(rcols, data_rows, final_idx):
        """Std (ddof=0) of the numeric per-rep selectivities at the final time."""
        vals = [data_rows[final_idx][c] for c in rcols]
        vals = [float(v) for v in vals if isinstance(v, (int, float))]
        if len(vals) > 1:
            mu = sum(vals) / len(vals)
            return (sum((x - mu) ** 2 for x in vals) / len(vals)) ** 0.5
        return 0.0

    def split_name(n):
        m = re.match(r"\s*(\S+wt%)\s*(.*)$", n)
        if m:
            return m.group(1), m.group(2).strip()
        return n, ""

    out = []
    if mode == "current_density":
        # one figure per wt fraction: selectivity vs current density at final time
        names, mean_cols, rep_col_lists, data_rows, times, final_idx = \
            parse_section("CON-based selectivity")
        from collections import OrderedDict as OD
        bywt = OD()
        for n, mc, rc in zip(names, mean_cols, rep_col_lists):
            wt, cur = split_name(n)
            cd = parse_current_density(cur, None)
            sval = data_rows[final_idx][mc]
            if isinstance(sval, (int, float)):
                bywt.setdefault(wt, []).append((cd, float(sval), rep_std(rc, data_rows, final_idx)))
        for wi, (wt, pts) in enumerate(bywt.items()):
            pts = sorted(pts)
            x = [p[0] for p in pts]
            y = [p[1] for p in pts]
            e = [p[2] for p in pts]
            fig, ax = plt.subplots(figsize=(3.6, 3.0))
            ax.errorbar(x, y, yerr=e, fmt="o-", color=_NPG[wi % len(_NPG)], ms=6, lw=1.4,
                        capsize=3, capthick=0.9, elinewidth=0.9)
            ax.set_xlabel(r"Current density (mA cm$^{-2}$)")
            ax.set_ylabel(r"Permselectivity $S_{\mathrm{Ca/Na}}$ (CON)")
            ax.set_title(f"{wt}  (t = {times[final_idx]} min)", fontsize=9)
            ax.grid(True, alpha=0.3, ls="--")
            fig.tight_layout()
            tag = wt.replace("%", "").replace(" ", "")
            for fmt in ("png", "pdf"):
                p = os.path.join(str(out_dir), f"Figure3_Selectivity_{tag}.{fmt}")
                fig.savefig(p, format=fmt, bbox_inches="tight")
                out.append(p)
            plt.close(fig)
    else:
        # doping mode: one bar figure per selectivity basis (CON- and DIL-based).
        # x-axis = resin loading, numeric wt% ascending (LOW->HIGH); any non-wt%
        # reference group (e.g. a commercial membrane) is appended at the right
        # in original order and drawn in grey.  CON keeps the original filename;
        # DIL is written as Figure3_Selectivity_DIL.*
        def _loading(n):
            m = re.search(r"(\d+(?:\.\d+)?)\s*wt%", str(n))
            return float(m.group(1)) if m else float("nan")
        for section_label, suffix, basis in (
                ("CON-based selectivity", "", "CON"),
                ("DIL-based selectivity", "_DIL", "DIL")):
            if basis not in doping_bases:
                continue
            try:
                names, mean_cols, rep_col_lists, data_rows, times, final_idx = \
                    parse_section(section_label)
            except StopIteration:
                continue
            rows_d = []
            for n, mc, rc in zip(names, mean_cols, rep_col_lists):
                sval = data_rows[final_idx][mc]
                rows_d.append((_loading(n), str(n),
                               float(sval) if isinstance(sval, (int, float)) else float("nan"),
                               rep_std(rc, data_rows, final_idx)))
            numeric = sorted([r for r in rows_d if not math.isnan(r[0])], key=lambda x: x[0])
            nonnum = [r for r in rows_d if math.isnan(r[0])]
            rows_d = numeric + nonnum
            n_num = len(numeric)
            labels = [(f"{r[0]:g}" if not math.isnan(r[0]) else r[1]) for r in rows_d]
            yvals = [r[2] for r in rows_d]
            yerr = [r[3] for r in rows_d]
            # sequential blue shades deepen with loading; grey for non-wt% refs
            nb = len(labels)
            bar_colors = []
            for i in range(nb):
                if i < n_num:
                    t = 0.30 + 0.62 * i / (n_num - 1) if n_num > 1 else 0.6
                    bar_colors.append(plt.cm.Blues(t))
                else:
                    bar_colors.append("#9E9E9E")
            fig, ax = plt.subplots(figsize=(5.2, 3.2))
            ax.bar(range(nb), yvals, yerr=yerr, capsize=3,
                   error_kw={"elinewidth": 0.9, "capthick": 0.9},
                   color=bar_colors, edgecolor="#333333", linewidth=0.4)
            ax.set_xticks(range(nb))
            ax.set_xticklabels(labels)
            ax.set_xlabel("Resin loading (wt%)")
            ax.set_ylabel(r"Permselectivity $S_{\mathrm{Ca/Na}}$ (%s)" % basis)
            ax.set_title(f"t = {times[final_idx]} min", fontsize=9)
            ax.grid(True, axis="y", alpha=0.3, ls="--")
            fig.tight_layout()
            for fmt in ("png", "pdf"):
                p = os.path.join(str(out_dir), f"Figure3_Selectivity{suffix}.{fmt}")
                fig.savefig(p, format=fmt, bbox_inches="tight")
                out.append(p)
            plt.close(fig)
    return out


def plot_selectivity_timecourse(xlsx_path, out_dir, vs_overrides=None, exclude=None,
                                fname="Figure3_Selectivity_timecourse"):
    """Time-resolved CON- and DIL-based selectivity vs time (two stacked panels).

    Recomputes per-replicate selectivity at EVERY measured time straight from
    ``1_Original Data`` (so it is independent of the endpoint export in sheets
    2-8).  Each group is one line: mean over valid reps, error bar = replicate
    std (ddof=0); NQ points (JCa_eff<=0 / JNa_eff<=0 / |JNa_eff|<1e-6) are
    dropped.  Top panel = CON-based, bottom = DIL-based (DIL uses the removed
    amount, -J).  ``vs_overrides`` {(group,rep):V_sample_mL} as in
    rebuild_results_workbook.
    """
    import os
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams.update({
        "font.family": "sans-serif", "font.sans-serif": ["Arial", "DejaVu Sans"],
        "font.size": 8, "axes.linewidth": 0.8, "pdf.fonttype": 42, "ps.fonttype": 42,
        "savefig.dpi": 600,
    })
    _PAL = ["#E64B35", "#4DBBD5", "#00A087", "#3C5488", "#F39B7F", "#8491B4",
            "#91D1C2", "#DC0000", "#7E6148"]
    _MK = ["o", "s", "^", "D", "v", "p", "X", "h", "*"]

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    data = read_original(wb["1_Original Data"])
    exclude = set(exclude or ())
    for group, side_dict in data.items():
        for side in ("CON", "DIL"):
            for rep, tv in side_dict[side].items():
                for t in list(tv):
                    if (group, t) in exclude:
                        tv[t] = (None, None)
    vs_overrides = dict(vs_overrides or {})
    groups = list(data.keys())
    all_times = sorted({t for g in data.values() for side in ("CON", "DIL")
                        for rep in g[side].values() for t in rep})

    def _rep_std(per_rep):
        vals = [v for v in per_rep.values() if isinstance(v, (int, float))]
        if len(vals) > 1:
            mu = sum(vals) / len(vals)
            return (sum((x - mu) ** 2 for x in vals) / len(vals)) ** 0.5
        return 0.0

    sel = {}
    for n in groups:
        sel[n] = {}
        for side, removal in (("CON", False), ("DIL", True)):
            reps_here = sorted(data[n][side].keys())
            vs_map = {rep: vs_overrides.get((n, rep), V_sample_mL) for rep in reps_here}
            comp = compute_group(data[n][side], all_times, all_times, vs_map)
            srows = _selectivity_rows(comp["reps"], comp["jca"], comp["jna"], all_times,
                                      removal=removal)
            sel[n][side] = {t: (srows[t]["mean"], _rep_std(srows[t]["per_rep"]))
                            for t in all_times if t != 0}

    fig, (axc, axd) = plt.subplots(2, 1, figsize=(5.6, 6.4), sharex=True)
    for axi, side, title in ((axc, "CON", "CON-based"), (axd, "DIL", "DIL-based")):
        for gi, n in enumerate(groups):
            d = sel[n][side]
            ts = [t for t in all_times if t != 0 and d.get(t) and d[t][0] is not None]
            if not ts:
                continue
            ys = [d[t][0] for t in ts]
            es = [d[t][1] for t in ts]
            axi.errorbar(ts, ys, yerr=es, fmt=_MK[gi % len(_MK)] + "--",
                         color=_PAL[gi % len(_PAL)], ms=5, lw=1.2, capsize=2.5,
                         capthick=0.8, elinewidth=0.8, label=n)
        axi.axhline(1.0, color="#888888", lw=0.7, ls=":")
        axi.set_ylabel(r"Permselectivity $S_{\mathrm{Ca/Na}}$ (%s)" % side)
        axi.set_title(title, fontsize=9)
    axd.set_xlabel("Time (min)")
    axc.legend(loc="best", fontsize=6.5, ncol=2)
    fig.tight_layout()
    out = []
    for fmt in ("png", "pdf"):
        p = os.path.join(str(out_dir), f"{fname}.{fmt}")
        fig.savefig(p, format=fmt, bbox_inches="tight")
        out.append(p)
    plt.close(fig)
    return out
