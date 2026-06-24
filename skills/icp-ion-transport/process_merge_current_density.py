# -*- coding: utf-8 -*-
"""
Merge the two parallel current-density runs into one ``1_Original Data`` table.

Sources (cleaned raw ICP exports, staged in ../../Original):
  * Na, Ca_ 20260522.csv  -> 5wt% & 10wt% x {1,5,10 mA*cm^-2}   (reps 1,2)
  * Na, Ca_ 20260529.csv  -> 5wt% x {1,5,10 mA*cm^-2} REPEAT     (reps -> 3,4)

The 5wt% groups therefore pool 4 replicate cells (run-1 reps 1/2 + run-2 reps
3/4); 10wt% groups keep 2 replicates (run-1 only).

Calcium is read on **Ca 317.933** (the only Ca line measured in the repeat run,
also present in run-1) so both runs share one wavelength.  Per-block piecewise
linear calibration, dilution x75, identical to the established pipeline.

Cleaning applied here (logged to ../../../CLEANING_REPORT.md):
  * only canonical labels  "{wt}wt% -{t}-{CON|DIL}-{rep} {I}mA*cm^-2"  kept;
    the tolerant regex absorbs missing hyphens / double spaces.
  * trailing standalone D0..D180 / C0..C180 series in 20260522 are dropped.

Writes Results.xlsx (sheets: 1_Original Data, Calibration).  Run
calculate_results.py afterwards to add sheets 2-8.
"""
import csv
import os
import re
from collections import OrderedDict

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from scipy.optimize import curve_fit
from scipy.stats import linregress

HERE = os.path.dirname(os.path.abspath(__file__))
ORIG = os.path.abspath(os.path.join(HERE, "..", "..", "Original"))

DILUTION = 75
CA_WL = "Ca 317.933"
NA_WL = "Na 588.995"
TARGET_WLS = OrderedDict([(CA_WL, "Ca"), (NA_WL, "Na")])
BREAK_GUESSES = {CA_WL: 40.0, NA_WL: 40.0}

# files: (filename, rep_offset)  -- second run's reps shifted by +2 to pool with run-1
DATA_FILES = [
    ("Na, Ca_ 20260522.csv", 0),
    ("Na, Ca_ 20260529.csv", 2),
]
# Na STD for a group's 1mA block is weak -> pool from that run's 5/10 mA blocks
NA_CAL_OVERRIDES = {"5wt% 1mA*cm^-2": ["5wt% 5mA*cm^-2", "5wt% 10mA*cm^-2"]}

# Data-check decision (2026-06-16): the 1 mA*cm^-2 groups are diffusion-dominated
# (DIL flux positive, current efficiency >200%, transport ~ noise floor), so they
# are removed entirely and leave no record in 1_Original Data.
DROP_GROUPS = {"5wt% 1mA*cm^-2", "10wt% 1mA*cm^-2"}


def to_float(value):
    if value is None:
        return None
    s = str(value).strip()
    if s in ("####", "Uncal", "-", "", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_measurements(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader); next(reader)
        header = next(reader)
        rows = [dict(zip(header, r)) for r in reader if len(r) >= len(header)]
    measurements, cur_key, cur = [], None, []
    for row in rows:
        key = (row.get("Label", ""), row.get("Type", ""), row.get("Date Time", ""))
        if cur_key is None:
            cur_key = key
        if key != cur_key:
            measurements.append(cur); cur = []; cur_key = key
        cur.append(row)
    if cur:
        measurements.append(cur)
    return measurements


def split_blocks(measurements):
    blocks, cur = [], None
    for meas in measurements:
        typ = meas[0].get("Type", "").strip()
        if typ == "BLK":
            if cur is not None:
                blocks.append(cur)
            cur = {"std": [meas], "samples": []}
        elif typ == "STD":
            if cur is not None:
                cur["std"].append(meas)
        elif typ == "Sample":
            if cur is not None:
                cur["samples"].append(meas)
    if cur is not None:
        blocks.append(cur)
    return blocks


def get_value(meas, element, field):
    for row in meas:
        if row.get("Element") == element:
            return to_float(row.get(field))
    return None


SAMPLE_RE = re.compile(
    r"^(?P<wt>\d+wt%)\s*-?\s*(?P<time>\d+)\s*-\s*(?P<side>CON|DIL)\s*-\s*(?P<rep>\d+)\s+"
    r"(?P<current>\d+mA\*cm\^-2)\s*$"
)


def parse_label(label):
    m = SAMPLE_RE.match(str(label).strip())
    if not m:
        return None
    d = m.groupdict()
    return {"group": f"{d['wt']} {d['current']}", "side": d["side"],
            "rep": int(d["rep"]), "time": int(d["time"])}


def assign_block_groups(blocks):
    for blk in blocks:
        g = "Unparsed"
        for meas in blk["samples"]:
            p = parse_label(meas[0].get("Label", ""))
            if p:
                g = p["group"]; break
        blk["group"] = g


def piecewise_linear(x, xb, k1, b1, k2):
    b2 = k1 * xb + b1 - k2 * xb
    return np.where(x <= xb, k1 * x + b1, k2 * x + b2)


def r2_score(y, yhat):
    y = np.asarray(y, float); yhat = np.asarray(yhat, float)
    denom = np.sum((y - y.mean()) ** 2)
    return np.nan if denom == 0 else 1 - np.sum((y - yhat) ** 2) / denom


def fit_piecewise(conc, intensity, guess):
    conc = np.asarray(conc, float); intensity = np.asarray(intensity, float)
    order = np.argsort(conc); conc, intensity = conc[order], intensity[order]
    k0, b0, *_ = linregress(conc, intensity)
    popt, _ = curve_fit(piecewise_linear, conc, intensity,
                        p0=[guess, k0, b0, k0 * 0.75],
                        bounds=([conc[1], 0, -np.inf, 0], [conc[-2], np.inf, np.inf, np.inf]),
                        maxfev=20000)
    return popt


def make_cal(wl, conc, intensity, source):
    conc = np.asarray(conc, float); intensity = np.asarray(intensity, float)
    order = np.argsort(conc); conc, intensity = conc[order], intensity[order]
    popt = fit_piecewise(conc, intensity, BREAK_GUESSES[wl])
    xb, k1, b1, k2 = popt
    b2 = k1 * xb + b1 - k2 * xb
    yhat = piecewise_linear(conc, *popt)
    m1 = conc <= xb; m2 = conc >= xb
    return {"popt": popt, "b2": b2, "r2": r2_score(intensity, yhat),
            "r2_seg1": r2_score(intensity[m1], piecewise_linear(conc[m1], *popt)),
            "r2_seg2": r2_score(intensity[m2], piecewise_linear(conc[m2], *popt)),
            "conc": conc, "intensity": intensity, "source": source}


def calc_conc(cal, intensity):
    xb, k1, b1, k2 = cal["popt"]
    i_break = k1 * xb + b1
    iv = float(intensity)
    return (iv - b1) / k1 if iv <= i_break else (iv - cal["b2"]) / k2


def build_calibration(block):
    cal = {}
    for wl in TARGET_WLS:
        cl, il = [], []
        for meas in block["std"]:
            typ = meas[0].get("Type", "").strip()
            c = 0.0 if typ == "BLK" else get_value(meas, wl, "Concentration")
            inten = get_value(meas, wl, "Intensity")
            if c is None or inten is None:
                continue
            cl.append(c); il.append(inten)
        cal[wl] = make_cal(wl, cl, il, "own block") if len(cl) >= 3 else None
    return cal


def apply_na_overrides(blocks, cals):
    idx = {b["group"]: i for i, b in enumerate(blocks)}
    for tgt, srcs in NA_CAL_OVERRIDES.items():
        ti = idx.get(tgt)
        if ti is None:
            continue
        cl, il, used = [], [], []
        for s in srcs:
            si = idx.get(s)
            if si is None or cals[si].get(NA_WL) is None:
                continue
            cl += cals[si][NA_WL]["conc"].tolist()
            il += cals[si][NA_WL]["intensity"].tolist()
            used.append(s)
        if len(cl) >= 4:
            cals[ti][NA_WL] = make_cal(NA_WL, cl, il, "pooled from " + " + ".join(used))


def process_file(path, rep_offset):
    blocks = split_blocks(read_measurements(path))
    assign_block_groups(blocks)
    cals = [build_calibration(b) for b in blocks]
    apply_na_overrides(blocks, cals)
    results, cal_records, dropped = [], [], []
    for bi, (blk, cal) in enumerate(zip(blocks, cals), start=1):
        used = False
        for meas in blk["samples"]:
            label = meas[0].get("Label", "")
            p = parse_label(label)
            if not p:
                lab = str(label).strip()
                if lab:
                    dropped.append(lab)
                continue
            if p["group"] in DROP_GROUPS:   # intentionally removed (1 mA groups)
                continue
            ca_int = get_value(meas, CA_WL, "Intensity")
            na_int = get_value(meas, NA_WL, "Intensity")
            ca_ppm = calc_conc(cal[CA_WL], ca_int) * DILUTION if (ca_int is not None and cal.get(CA_WL)) else None
            na_ppm = calc_conc(cal[NA_WL], na_int) * DILUTION if (na_int is not None and cal.get(NA_WL)) else None
            results.append({
                "Group": p["group"], "Side": p["side"], "Rep": p["rep"] + rep_offset,
                "Time": p["time"], "Label": label, "Ca_int": ca_int, "Na_int": na_int,
                "Ca_ppm": round(ca_ppm, 4) if ca_ppm is not None else None,
                "Na_ppm": round(na_ppm, 4) if na_ppm is not None else None,
            })
            used = True
        if used:
            for wl in TARGET_WLS:
                cal_w = cal.get(wl)
                if cal_w is None:
                    continue
                xb, k1, b1, k2 = cal_w["popt"]
                cal_records.append([os.path.basename(path), bi, blk["group"], wl, "piecewise_linear",
                                    len(cal_w["conc"]), f"{cal_w['conc'].min():.0f}-{cal_w['conc'].max():.0f}",
                                    xb, k1, b1, k2, cal_w["b2"], cal_w["r2"], cal_w["r2_seg1"],
                                    cal_w["r2_seg2"], cal_w["source"]])
    return results, cal_records, dropped


# ---- group/rep ordering ---------------------------------------------------
GROUP_ORDER = ["5wt% 1mA*cm^-2", "5wt% 5mA*cm^-2", "5wt% 10mA*cm^-2",
               "10wt% 1mA*cm^-2", "10wt% 5mA*cm^-2", "10wt% 10mA*cm^-2"]


def main():
    all_results, all_cal, all_dropped = [], [], []
    for fname, off in DATA_FILES:
        res, cal, dropped = process_file(os.path.join(ORIG, fname), off)
        all_results += res; all_cal += cal; all_dropped += [(fname, d) for d in dropped]
        print(f"{fname}: {len(res)} samples, {len(set(d for _,d in [(fname,x) for x in dropped]))} distinct dropped labels")

    # sort: group order, side (CON before DIL), rep, time
    side_rank = {"CON": 0, "DIL": 1}
    all_results.sort(key=lambda r: (GROUP_ORDER.index(r["Group"]) if r["Group"] in GROUP_ORDER else 99,
                                    side_rank.get(r["Side"], 9), r["Rep"], r["Time"]))

    wb = Workbook()
    ws = wb.active; ws.title = "1_Original Data"
    headers = ["Group", "Side", "Rep", "Time (min)", "Label",
               f"{CA_WL} Intensity", "Ca (ppm, x75)", f"{NA_WL} Intensity", "Na (ppm, x75)"]
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Border(*(Side("thin"),) * 4)
    ctr = Alignment(horizontal="center", vertical="center")
    palette = ["DAEEF3", "E2EFDA", "FFF2CC", "FCE4D6", "E4DFEC", "D9EAD3"]
    gcolors = {}
    for r in all_results:
        gcolors.setdefault(r["Group"], palette[len(gcolors) % len(palette)])
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = ctr; cell.border = thin
    for ri, r in enumerate(all_results, 2):
        vals = [r["Group"], r["Side"], r["Rep"], r["Time"], r["Label"],
                r["Ca_int"], r["Ca_ppm"], r["Na_int"], r["Na_ppm"]]
        fill = PatternFill("solid", fgColor=gcolors[r["Group"]])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(ri, c, v); cell.fill = fill; cell.border = thin; cell.alignment = ctr
            if c in (6, 8):
                cell.number_format = "0.00"
            elif c in (7, 9):
                cell.number_format = "0.0000"
    for i, w in enumerate([22, 8, 6, 12, 30, 20, 16, 20, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    wc = wb.create_sheet("Calibration")
    chead = ["Source", "Block", "Group", "Wavelength", "Fit", "Points", "Range (ppm)", "Breakpoint ppm",
             "k1", "b1", "k2", "b2", "R2", "Seg1 R2", "Seg2 R2", "Source detail"]
    for c, h in enumerate(chead, 1):
        cell = wc.cell(1, c, h); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = thin
    for ri, rec in enumerate(all_cal, 2):
        for c, v in enumerate(rec, 1):
            cell = wc.cell(ri, c, v); cell.border = thin; cell.alignment = ctr
            if 8 <= c <= 15:
                cell.number_format = "0.00000000"

    out = os.path.join(HERE, "Results.xlsx")
    wb.save(out)
    # report counts
    counts = OrderedDict()
    for r in all_results:
        counts.setdefault(r["Group"], set()).add((r["Side"], r["Rep"]))
    print(f"\nSaved: {out}")
    for g in GROUP_ORDER:
        if g in counts:
            reps = sorted(set(rep for s, rep in counts[g]))
            print(f"  {g}: reps {reps}")
    print(f"  dropped (non-canonical) labels total: {len(all_dropped)}")


if __name__ == "__main__":
    main()
